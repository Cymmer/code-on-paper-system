import requests
from openpyxl import load_workbook
import jsmin

import xlwt
import xlrd
from xlutils.copy import copy
import editdistance

from analysis import minify_c_code, string_similarity
from python_minifier import minify
import re

problem_number = 9
file_path = f"/Users/cymmerjohnmaranga/Downloads/Manipulated/C_Problem_{problem_number}_Test.xlsx"
workbook = load_workbook(filename=file_path)
sheet = workbook.active

start_col = 3
end_col = 4
minified_data = []
for i in range(2, sheet.max_row + 1):
    row_data = [cell.value for cell in sheet[i][start_col : end_col + 1]]
    row_data_minified = []
    for col in row_data:
        row_data_minified.append(minify_c_code(col))
    minified_data.append(row_data_minified)

# load the excel file
rb = xlrd.open_workbook(file_path)

# copy the contents of excel file
wb = copy(rb)

# open the first sheet
w_sheet = wb.get_sheet(0)
w_sheet_by_index = rb.sheet_by_index(0)

for i, data_row in enumerate(minified_data, start=1):
    print(data_row)

    corrected_code_minified = minify_c_code(data_row[0])
    actual_code_minified = minify_c_code(data_row[1])
    google_vision_minified = w_sheet_by_index.cell(i, start_col+2).value

    # Open AI Minified
    w_sheet.write(i, start_col+3, corrected_code_minified)
    # Actual Code Minified
    w_sheet.write(i, start_col+4, actual_code_minified)

    # Google Vision API vs Corrected Code
    w_sheet.write(i, start_col+5, string_similarity(google_vision_minified, corrected_code_minified))

    # Corrected Code vs Actual Code
    w_sheet.write(i, start_col+6, string_similarity(corrected_code_minified, actual_code_minified))


# save the file
wb.save(f"/Users/cymmerjohnmaranga/Downloads/Manipulated/Minified/C_Problem_{problem_number}_Test_with_minified.xls")

from openpyxl import load_workbook
import re


def minify_c_code(c_code):
    # Remove comments
    c_code = re.sub(r'\/\/.*', '', c_code)
    c_code = re.sub(r'\/\*[\s\S]*?\*\/', '', c_code)

    # Remove newlines
    c_code = re.sub(r'\n', '', c_code)

    # Minify spaces within print statements
    c_code = re.sub(r'(\bprintf\("[^"]*)\s+(\S*"\))', r'\1 \2', c_code)

    # Remove remaining spaces and semicolons
    c_code = re.sub(r'\s+', ' ', c_code)
    c_code = c_code.strip().rstrip(';')

    return c_code

file_path = "/Users/cymmerjohnmaranga/Downloads/C_Problem3_Dataset-1.xlsx"
workbook = load_workbook(filename=file_path)
sheet = workbook.active

start_col = 4
end_col = 4
minified_data = []
# for i in range(2, sheet.max_row + 1):
#     row_data = [cell.value for cell in sheet[i][start_col : end_col + 1]]
#     row_data_minified = []
#     for col in row_data:
#         row_data_minified.append(minify_c_code(col))
#     minified_data.append(row_data_minified)

for i in range(2, sheet.max_row + 1):
    row_data = [cell.value for cell in sheet[i][start_col : end_col + 1]]
    minified_data.append(len(minify_c_code(row_data[0]).strip()))

print(sum(minified_data)/len(minified_data))

# load the excel file
# rb = xlrd.open_workbook("/Users/cymmerjohnmaranga/Downloads/C_Problem1_Dataset-1.xlsx")

# # copy the contents of excel file
# wb = copy(rb)

# # open the first sheet
# w_sheet = wb.get_sheet(0)

# for i, data_row in enumerate(minified_data, start=1):
#     actual_results = None

#     for j, data in enumerate(data_row, start=6):
#         # print(j)
#         if j == 8: actual_results = data
#         w_sheet.write(i, j, data)

#     # Exclude actual results
#     for j, data in enumerate(data_row[:2], start=2):
#         # print(data, "#########", actual_results)
#         # print("",end="\n\n\n\n\n")
#         # break
#         if actual_results:
#             distance = editdistance.eval(data, actual_results)
#             percentage = (1 - distance / max(len(data), len(actual_results))) * 100

#             w_sheet.write(i, j+7, percentage)

# # save the file
# wb.save("/Users/cymmerjohnmaranga/Downloads/C_Problem1_Dataset-1-with-minified.xls")
